from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import requests
import json
from jira import JIRA
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill
import itertools as it

class JiraATT(object):
	'''jira class using an attuid and password'''

	def __init__(self, attuid, password):
		'''setup user's credentials and paths, rsa file path, username, and password'''
		##############################################################################
		self.base_url = '' # the base url to browse a single ticket
		self.login_url = self.base_url+'/login.jsp' # the login page
		self.ticket_base = self.base_url+'/browse' # the ticket page
		self.filter_api = self.base_url+'/rest/api/2/filter'
		##############################################################################
		self.dividers = 234 # how long are the dividers in the ascii table
		self.total_tickets = 0
		##############################################################################
		self.attuid = attuid
		self.password = password
		###############################################################################
		self.qa_begin = "h2. ============================ QA Steps ============================"
		self.qa_end = "h2. ================================================================="
		self.qa_regex_begin = re.compile(r"h2\. ============================ QA Steps ============================")
		self.qa_regex_end = re.compile(r"h2\. =================================================================")
		###############################################################################
	
	def get_ticket_base(self):
		return self.ticket_base

	def login(self, attuid='', password=''):
		if attuid and password:
			self.attuid = attuid
			self.password = password
		'''create a session and logs into jira'''
		self.session_obj = JIRA( self.base_url, basic_auth=(self.attuid, self.password) )

	def reset_data(self):
		self.keys = []
		self.msrps = []
		self.statuses = []
		self.components = []
		self.summaries = []
		self.story_points = []
		self.attuids = []
		self.sprints = []

		self.customer_names = []
		self.customer_attuids = []
		self.team_orgs = []
		self.epic_links = []
		self.resolutions = []
		self.priorities = []
		self.issue_types = []
		self.versions = []
		self.severities = []

		self.due_dates = []
		self.create_dates = []
		self.updated_dates = []
		self.start_dates = []
		self.labels = []

		self.qa_steps = []
		self.comments = []


	def get_issues(self, filter_number):
		self.issues = self.session_obj.search_issues('filter='+filter_number, maxResults=500, fields='comment,key,customfield_10212,status,summary,assignee,components,customfield_10006,customfield_10001,customfield_10103,customfield_10102,customfield_10002,customfield_10101,resolution,priority,issuetype,fixVersions,customfield_10108,duedate,created,updated,customfield_10109,comment,labels')
		print('found', len(self.issues),'issues.')

	def get_jira_data(self, filter_number):
		self.reset_data()
		self.total_tickets = 0
		self.get_issues(filter_number)
		self.parse_data()

	def parse_data(self):
		# get data
		for idx, issue in enumerate(self.issues):

			# print all issue properties
			# print(', '.join("%s: %s" % item for item in vars(issue.fields).items()))

			# get key
			self.keys.append(issue.key)
			# get msrp
			self.msrps.append(issue.fields.customfield_10212)
			# get status
			self.statuses.append(issue.fields.status.name)
			# get summary
			self.summaries.append(issue.fields.summary)

			# get assignee
			if issue.fields.assignee:
				self.attuids.append( self.generate_attuid(issue.fields.assignee.name) )
			else:
				self.attuids.append('')

			# get component
			all_components = []
			if issue.fields.components:
				for i in range(len(issue.fields.components)):
					all_components.append(str(issue.fields.components[i]))
			else:
				all_components.append('')
			self.components.append(' '.join(all_components))

			# get story points
			if issue.fields.customfield_10006:
				self.story_points.append(issue.fields.customfield_10006)
			else:
				self.story_points.append(0)

			# get sprint
			sprint = ''
			if issue.fields.customfield_10001:
				sprint_split = issue.fields.customfield_10001[0].split(',')
				if len(sprint_split) > 3:
					sprint_split = sprint_split[3].split('=')
					if len(sprint_split) > 1:
						sprint = sprint_split[1]
			self.sprints.append(sprint)

			# get customer name
			if issue.fields.customfield_10103:
				self.customer_names.append(issue.fields.customfield_10103)
			else:
				self.customer_names.append('')

			# get customer attuid
			if issue.fields.customfield_10102:
				self.customer_attuids.append(issue.fields.customfield_10102)
			else:
				self.customer_attuids.append('')

			# get team org 
			if issue.fields.customfield_10101:
				self.team_orgs.append(issue.fields.customfield_10101)
			else:
				self.team_orgs.append('')

			# get epic link
			epic_link = issue.fields.customfield_10002
			if epic_link == 'UD-2421':
				epic_link = 'Apollo'
			elif epic_link == 'UD-1':
				epic_link = 'Gamma'
			elif epic_link == 'UD-3532':
				epic_link = 'Ember Upgrades'
			elif epic_link == 'UD-3':
				epic_link = 'Magellan'
			elif epic_link == 'UD-4714':
				epic_link = 'UTM'
			elif epic_link == 'UD-656':
				epic_link = 'US GCSC'
			else:
				epic_link = ''
			self.epic_links.append(epic_link)

			# get resolution
			if issue.fields.resolution:
				self.resolutions.append(issue.fields.resolution)
			else:
				self.resolutions.append('')

			# get priority
			if issue.fields.priority:
				self.priorities.append(issue.fields.priority)
			else:
				self.priorities.append('')

			# get issue type
			if issue.fields.issuetype:
				self.issue_types.append(issue.fields.issuetype)
			else:
				self.issue_types.append('')

			# get versions
			if issue.fields.fixVersions:
				self.versions.append(issue.fields.fixVersions)
			else:
				self.versions.append('')

			# get severities
			if issue.fields.customfield_10108:
				self.severities.append(issue.fields.customfield_10108)
			else:
				self.severities.append('')

			# get due dates
			if issue.fields.duedate:
				self.due_dates.append(issue.fields.duedate)
			else:
				self.due_dates.append('')

			# get create dates
			if issue.fields.created:
				self.create_dates.append(issue.fields.created)
			else:
				self.create_dates.append('')

			# get updated dates
			if issue.fields.updated:
				self.updated_dates.append(issue.fields.updated)
			else:
				self.updated_dates.append('')

			# get start dates
			if issue.fields.customfield_10109:
				self.start_dates.append(issue.fields.customfield_10109)
			else:
				self.start_dates.append('')

			# get start dates
			if issue.fields.labels:
				self.labels.append(issue.fields.labels)
			else:
				self.labels.append('')

			# get comments and QA steps
			if issue.fields.comment:
				comments = []
				issue_qa_step = ''
				missing_qa_step = ''
				found_qa = False;
				# for each comment save it and see if QA steps
				for comment in issue.fields.comment.comments:
					# save comment
					comments.append(comment.body)
					# if in code review status then QA steps must exist
					if (issue.fields.status.name.lower() in ['code review', 'ready for qa', 'ready for uct']):
						# test if QA steps needed
						result = re.split(self.qa_regex_begin, str(comment.body))
						# if we found beginning of QA then let's find end
						if len(result) > 1 and not found_qa:
							result = re.split(self.qa_regex_end, str(result[1]))
							# if found end of QA then we've found QA
							if len(result) > 1:
								issue_qa_step = result[0]
								found_qa = True
							else:
								missing_qa_step = 'Missing end marker of QA steps'
						elif not found_qa:
							missing_qa_step = 'Missing begin marker of QA steps'

				# save QA step and comments
				if not issue_qa_step:
					issue_qa_step = missing_qa_step
				self.qa_steps.append(issue_qa_step)
				self.comments.append(comments)
			else:
				self.qa_steps.append('')
				self.comments.append([])

		# get total ticket
		self.total_tickets = len(self.keys)
		self._check_ticket_fields()
		

	def show_jira_ascii_header(self):
		# {0:15.15} = use variable 0 ('Key') with 15 preallocated spaces, .15 means slice string after 15 characters
		print('', '-'* self.dividers)
		print(" | {0:10} | {1:7} | {2:20} | {3:22} | {4:65} | {5:52} | {6:6} | {7:6} | {8:8} | {9:7} |".format('Key', 'MSRP', 'Status', 'Component', 'Summary', 'URL', 'Attuid', 'Points', 'Labels', 'Epic'))
		print('', '-'* self.dividers)

	def show_jira_ascii_footer(self):
		'''print ascii footer'''
		print('','-'* self.dividers)
		print(" | TOTAL: {0:3.3} ".format(str(self.total_tickets)), " "*(self.dividers-11), "|")
		print('','-'* self.dividers, '\n\n\n')

	def show_jira_ascii(self):
		'''print ascii of jira data to console, boolean to print total at bottom or not'''
		self.show_jira_ascii_header()
		for idx, val in enumerate(self.keys):
			# see if beta ticket
			is_beta = ''
			for label in self.labels[idx]:
				if label in ['BETA', 'NewHire', 'Blitz547', 'Blitz226', 'ASE']:
					is_beta = label
			# print ticket ascii
			print( " | {0:10.10} | {1:7.7} | {2:20.20} | {3:22.22} | {4:65.65} | {5}/{6:15.15} | {7:6.6} | {8:6.6} | {9:8.8} | {10:7.7} |".format(self.keys[idx], self.msrps[idx], self.statuses[idx], self.components[idx], self.summaries[idx], self.ticket_base, self.keys[idx], self.attuids[idx], str(self.story_points[idx]), is_beta, self.epic_links[idx]))
			if(idx % 4 == 0):
				print('','-'* self.dividers)
		self.show_jira_ascii_footer()

	def get_total_tickets(self):
		'''returns number of ticket in filter'''
		return self.total_tickets

	def login_api(self):
		'''create a session and logs into jira'''
		self.session_obj = requests.Session()
		self.session_obj.post(self.login_url, data=self.payload)

	def generate_attuid(self, attuid):
		'''get real attuid if not working'''
		if(len(attuid) == 6):
			return attuid
		else:
			return attuid[-6:]

	def _check_ticket_fields(self):
		'''makes sure to set some kind of data to field if it does not exist'''
		for idx, val in enumerate(self.keys):
			if not self.keys[idx]:
				self.keys[idx] = ''
			if not self.msrps[idx]:
				self.msrps[idx] = ''
			if not self.statuses[idx]:
				self.statuses[idx] = ''
			if not self.components[idx]:
				self.components[idx] = ''
			if not self.summaries[idx]:
				self.summaries[idx] = ''
			if not self.attuids[idx]:
				self.attuids[idx] = ''
			if not self.story_points[idx]:
				self.story_points[idx] = 0

	def sort_data_by(self, filter_key):
		all_data = []
		# get all data into dictionary
		for idx, val in enumerate(self.keys):
			all_data_dict = {
				'key': self.keys[idx],
				'msrp': self.msrps[idx],
				'status': self.statuses[idx],
				'component': self.components[idx],
				'summary': self.summaries[idx],
				'attuid': self.attuids[idx],
				'story_points': self.story_points[idx],
				'sprints': self.sprints[idx],
				'customer_names': self.customer_names[idx],
				'customer_attuids':self.customer_attuids[idx],
				'team_orgs': self.team_orgs[idx],
				'epic_links': self.epic_links[idx],
				'resolutions': self.resolutions[idx],
				'priorities': self.priorities[idx],
				'issue_types': self.issue_types[idx],
				'versions': self.versions[idx],
				'severities': self.severities[idx],
				'due_dates': self.due_dates[idx],
				'create_dates': self.create_dates[idx],
				'updated_dates': self.updated_dates[idx],
				'start_dates': self.start_dates[idx],
				'labels': self.labels[idx],
				'qa_steps': self.qa_steps[idx],
				'comments': self.comments[idx],
			}
			all_data.append(all_data_dict)

		# sort data by attuid
		all_data_sorted = sorted(all_data, key=lambda k: k[filter_key])
		self.reset_data()

		# set all data back into instance's variables
		for idx, val in enumerate(all_data_sorted):
			self.keys.append(all_data_sorted[idx]['key'])
			self.msrps.append(all_data_sorted[idx]['msrp'])
			self.statuses.append(all_data_sorted[idx]['status'])
			self.components.append(all_data_sorted[idx]['component'])
			self.summaries.append(all_data_sorted[idx]['summary'])
			self.attuids.append(all_data_sorted[idx]['attuid'])
			self.story_points.append(all_data_sorted[idx]['story_points'])
			self.sprints.append(all_data_sorted[idx]['sprints'])
			self.customer_names.append(all_data_sorted[idx]['customer_names'])
			self.customer_attuids.append(all_data_sorted[idx]['customer_attuids'])
			self.team_orgs.append(all_data_sorted[idx]['team_orgs'])
			self.epic_links.append(all_data_sorted[idx]['epic_links'])
			self.resolutions.append(all_data_sorted[idx]['resolutions'])
			self.priorities.append(all_data_sorted[idx]['priorities'])
			self.issue_types.append(all_data_sorted[idx]['issue_types'])
			self.versions.append(all_data_sorted[idx]['versions'])
			self.severities.append(all_data_sorted[idx]['severities'])
			self.due_dates.append(all_data_sorted[idx]['due_dates'])
			self.create_dates.append(all_data_sorted[idx]['create_dates'])
			self.updated_dates.append(all_data_sorted[idx]['updated_dates'])
			self.start_dates.append(all_data_sorted[idx]['start_dates'])
			self.labels.append(all_data_sorted[idx]['labels'])
			self.qa_steps.append(all_data_sorted[idx]['qa_steps'])
			self.comments.append(all_data_sorted[idx]['comments'])

	def return_jira_data(self):
		'''return all data from jira'''
		data = {
			'attuids': self.attuids,
			'keys': self.keys,
			'msrps': self.msrps,
			'statuses': self.statuses,
			'components': self.components,
			'summaries': self.summaries,
			'story_points': self.story_points,
			'sprints': self.sprints,
			'customer_names': self.customer_names,
			'customer_attuids':self.customer_attuids,
			'team_orgs': self.team_orgs,
			'epic_links': self.epic_links,
			'resolutions': self.resolutions,
			'priorities': self.priorities,
			'issue_types': self.issue_types,
			'versions': self.versions,
			'severities': self.severities,
			'due_dates': self.due_dates,
			'create_dates': self.create_dates,
			'updated_dates': self.updated_dates,
			'start_dates': self.start_dates,
			'labels': self.labels,
			'qa_steps': self.qa_steps,
			'comments': self.comments
		}
		return data

	def create_work_book(self):
		# create spreadsheet
		try:
			wb = Workbook()
			dest_filename = 'qa_steps.xlsx'
			ws = wb.active
			ws.page_setup.fitToHeight = 1
			ws.page_setup.fitToWidth = 1
			# create workbook headers
			ws.cell(column=1, row=1, value='MSRP Number')
			ws.cell(column=2, row=1, value='Status')
			ws.cell(column=3, row=1, value='Key')
			ws.cell(column=4, row=1, value='Summary')
			ws.cell(column=5, row=1, value='Epic Link')
			ws.cell(column=6, row=1, value='Sprint')
			ws.cell(column=7, row=1, value='Test Steps')
			# format headers
			for letter in 'ABCDEFG':
				ws[letter+'1'].font = Font(bold=True)
			# format cell sizes
			ws.column_dimensions['B'].width = 35
			ws.column_dimensions['C'].width = 15
			ws.column_dimensions['D'].width = 70
			ws.column_dimensions['E'].width = 15
			ws.column_dimensions['F'].width = 15
			ws.column_dimensions['G'].width = 100
			# for each issue if it has QA steps add to workbook
			workbook_index = 2
			for idx, issue in enumerate(self.keys):
				if(self.qa_steps[idx]):
					# format qa steps
					qa_steps = self.format_qa_steps(self.qa_steps[idx])
					# insert row data
					ws.cell(column=1, row=workbook_index, value=self.msrps[idx] )
					ws.cell(column=2, row=workbook_index, value=self.statuses[idx] )
					ws.cell(column=3, row=workbook_index, value='=HYPERLINK("{}", "{}")'.format(self.ticket_base+'/'+self.keys[idx], self.keys[idx]))
					ws.cell(column=4, row=workbook_index, value=self.summaries[idx] )
					ws.cell(column=5, row=workbook_index, value=self.epic_links[idx] )
					ws.cell(column=6, row=workbook_index, value=self.sprints[idx] )
					ws.cell(column=7, row=workbook_index, value=qa_steps )
					# format key links
					ws['C'+str(workbook_index)].font = Font(underline="single", color='FF0645AD')
					# got to next row
					workbook_index = workbook_index + 1    		
			# save workbook
			wb.save(filename = dest_filename)
			print('Workbook saved.')
		except Exception as e:
			print('Could not save workbook:',e)

	def format_qa_steps(self, qa_steps):
		cnt = it.count()
		next(cnt)
		qa_steps = qa_steps.replace('\=','').replace('\\','')
		qa_steps = re.sub(r"# ", lambda x: '{}. '.format(next(cnt)), qa_steps)
		return qa_steps

	def find_key_by_msrp(self, msrp):
		issue = self.session_obj.search_issues('MSRP_Number~'+msrp)
		if issue[0].key:
			return issue[0].key
		else:
			return False

	def find_qa_data(self, msrp):
		issue = self.session_obj.search_issues('MSRP_Number~'+msrp)
		if issue[0]:
			issue = issue[0]
			data = {}
			if issue.fields.customfield_10006:
				data['story_point'] = issue.fields.customfield_10006
			if issue.key:
				data['key'] = issue.key
			if issue.fields.summary:
				data['summary'] = issue.fields.summary
			return data
		else:
			return False

	def add_comment(self, key, comment, public_comment=False):
		response = ''
		if public_comment:
			response = self.session_obj.add_comment(key, comment)
		else:
			response = self.session_obj.add_comment(key, comment, visibility={'type': 'role', 'value': 'Developers'})
		return response

	def generate_qa_template(self, qa_steps, repos, crucible_id):
			repo_table = self.generate_repo_table(repos)
			return """
"""+repo_table+"""


"""+'[Crucible Review|https://icode3.web.att.com/cru/'+crucible_id+"""]


"""+self.qa_begin+"""

"""+qa_steps+"""

"""+self.qa_end+"""
			"""

	def generate_repo_table(self, repos):
		table_rows = []
		for repo in repos:
			baseBranch = repo['baseBranch']
			repositoryName = repo['repositoryName']
			reviewedBranch = repo['reviewedBranch']
			table_rows.append('|'+repositoryName+'|'+reviewedBranch+'|'+repositoryName+'|')

		table_header = "|| Repo || Branch || Branched From ||"
		table_data = ''
		for row in table_rows:
			table_data += """
"""+row
		return table_header + table_data