from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill


def create_work_book():
	'''
	Args:
		None
		
	Returns:
		None
	'''
	try:
		wb = Workbook()
		dest_filename = 'qa_steps.xlsx'
		ws = wb.active
		ws.page_setup.fitToHeight = 1
		ws.page_setup.fitToWidth = 1
		# create workbook headers
		ws.cell(column=1, row=1, value='MSRP Number')
		ws.cell(column=2, row=1, value='Status')
		ws.cell(column=3, row=1, value='Key')
		ws.cell(column=4, row=1, value='Summary')
		ws.cell(column=5, row=1, value='Epic Link')
		ws.cell(column=6, row=1, value='Sprint')
		ws.cell(column=7, row=1, value='Test Steps')
		# format headers
		for letter in 'ABCDEFG':
			ws[letter+'1'].font = Font(bold=True)
		# format cell sizes
		ws.column_dimensions['B'].width = 35
		ws.column_dimensions['C'].width = 15
		ws.column_dimensions['D'].width = 70
		ws.column_dimensions['E'].width = 15
		ws.column_dimensions['F'].width = 15
		ws.column_dimensions['G'].width = 100
		# for each issue if it has QA steps add to workbook
		workbook_index = 2
		for issue in issues:
			if(issue['qa_steps']):
				# format qa steps
				qa_steps = format_qa_steps(issue['qa_steps'])
				# insert row data
				ws.cell(column=1, row=workbook_index, value=issue['msrps'] )
				ws.cell(column=2, row=workbook_index, value=issue['statuses'] )
				ws.cell(column=3, row=workbook_index, value='=HYPERLINK("{}", "{}")'.format(ticket_base+'/'+issue['keys'], issue['keys']))
				ws.cell(column=4, row=workbook_index, value=issue['summaries'] )
				ws.cell(column=5, row=workbook_index, value=issue['epic_links'] )
				ws.cell(column=6, row=workbook_index, value=issue['sprints'] )
				ws.cell(column=7, row=workbook_index, value=issue['qa_steps'] )			# format key links
				ws['C'+str(workbook_index)].font = Font(underline="single", color='FF0645AD')
				# got to next row
				workbook_index += 1    		
		# save workbook
		wb.save(filename = dest_filename)
		print('Workbook saved.')
	except Exception as e:
		print('Could not save workbook:',e)